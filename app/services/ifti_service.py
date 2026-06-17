"""
IFTI (International Funds Transfer Instruction) Service.

Generates AUSTRAC-compatible Excel spreadsheets in the exact format
of IFTI-DRA IN and IFTI-DRA OUT official AUSTRAC templates.

Structure matches the official AUSTRAC spreadsheet:
  Row 0  — Section headers (with merged cell labels)
  Row 1  — Column sub-headers
  Row 2+ — One data row per transaction

Usage:
  workbook_bytes = generate_ifti_excel(records, direction="outgoing")
  # Save or stream as .xls / .xlsx
"""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.ifti import IFTIDirection, IFTIRecord

# ── Column definitions ────────────────────────────────────────────────────────
# Each entry: (section_label, column_label)
# Section label is shared across consecutive columns in that group.

IFTI_OUT_COLUMNS: list[tuple[str, str]] = [
    # Transaction details (7)
    ("Transaction details", "Date money/property received from the ordering customer"),
    (
        "Transaction details",
        "Date money/property made available to the beneficiary customer",
    ),
    ("Transaction details", "Currency code"),
    ("Transaction details", "Total amount/value"),
    ("Transaction details", "Type of transfer"),
    ("Transaction details", "Description of property"),
    ("Transaction details", "Transaction reference number"),
    # Ordering customer (3)
    ("Ordering customer", "Full name"),
    ("Ordering customer", "If known by any other name"),
    ("Ordering customer", "Date of birth (if an individual)"),
    # Ordering customer contact details (12)
    (
        "Ordering customer contact details",
        "Business/residential address (not a post box address)",
    ),
    ("Ordering customer contact details", "City/town/suburb"),
    ("Ordering customer contact details", "State"),
    ("Ordering customer contact details", "Postcode"),
    ("Ordering customer contact details", "Country"),
    ("Ordering customer contact details", "Postal address"),
    ("Ordering customer contact details", "City/town/suburb"),
    ("Ordering customer contact details", "State"),
    ("Ordering customer contact details", "Postcode"),
    ("Ordering customer contact details", "Country"),
    ("Ordering customer contact details", "Phone"),
    ("Ordering customer contact details", "Email"),
    # Ordering customer business details (5)
    (
        "Ordering customer business details",
        "Occupation, business or principal activity",
    ),
    ("Ordering customer business details", "ABN, ACN or ARBN"),
    ("Ordering customer business details", "Customer number (allocated by remitter)"),
    ("Ordering customer business details", "Account number (held by remitter)"),
    ("Ordering customer business details", "Business structure (if not an individual)"),
    # Ordering customer identification details (9)
    ("Ordering customer identification details", "ID type (1)"),
    ("Ordering customer identification details", "ID type (if 'Other')"),
    ("Ordering customer identification details", "Number"),
    ("Ordering customer identification details", "Issuer"),
    ("Ordering customer identification details", "ID type (2)"),
    ("Ordering customer identification details", "ID type (if 'Other')"),
    ("Ordering customer identification details", "Number"),
    ("Ordering customer identification details", "Issuer"),
    ("Ordering customer identification details", "Electronic data source"),
    # Beneficiary customer (3)
    ("Beneficiary customer", "Full name"),
    ("Beneficiary customer", "Date of birth (if an individual)"),
    (
        "Beneficiary customer",
        "Any business name under which the beneficiary customer is operating",
    ),
    # Beneficiary customer contact details (12)
    (
        "Beneficiary customer contact details",
        "Business/residential address (not a post box address)",
    ),
    ("Beneficiary customer contact details", "City/town/suburb"),
    ("Beneficiary customer contact details", "State"),
    ("Beneficiary customer contact details", "Postcode"),
    ("Beneficiary customer contact details", "Country"),
    ("Beneficiary customer contact details", "Postal address"),
    ("Beneficiary customer contact details", "City/town/suburb"),
    ("Beneficiary customer contact details", "State"),
    ("Beneficiary customer contact details", "Postcode"),
    ("Beneficiary customer contact details", "Country"),
    ("Beneficiary customer contact details", "Phone"),
    ("Beneficiary customer contact details", "Email"),
    # Beneficiary customer business details (2)
    (
        "Beneficiary customer business details",
        "Occupation, business or principal activity",
    ),
    ("Beneficiary customer business details", "ABN, ACN or ARBN"),
    (
        "Beneficiary customer business details",
        "Business structure (if not an individual)",
    ),
    # Beneficiary customer account details (4)
    ("Beneficiary customer account details", "Account number"),
    (
        "Beneficiary customer account details",
        "Name of institution (where account is held)",
    ),
    ("Beneficiary customer account details", "City"),
    ("Beneficiary customer account details", "Country"),
    # Person/org accepting transfer instruction (7+1 retail ID)
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Identification number of the retail outlet/business location",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Full name",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Business/residential address (not a post box address)",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "City/town/suburb",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "State",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Postcode",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Is this person/organisation accepting the money or property?",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Is this person/organisation sending the transfer instruction?",
    ),
    # Person/org accepting money (if different) (5)
    (
        "Person/organisation accepting the money or property from the ordering customer (if different)",
        "Full name",
    ),
    (
        "Person/organisation accepting the money or property from the ordering customer (if different)",
        "Business/residential address (not a post box address)",
    ),
    (
        "Person/organisation accepting the money or property from the ordering customer (if different)",
        "City/town/suburb",
    ),
    (
        "Person/organisation accepting the money or property from the ordering customer (if different)",
        "State",
    ),
    (
        "Person/organisation accepting the money or property from the ordering customer (if different)",
        "Postcode",
    ),
    # Person/org sending transfer instruction (if different) (16)
    (
        "Person/organisation sending the transfer instruction (if different)",
        "Full name",
    ),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "If known by any other name",
    ),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "Date of birth (if an individual)",
    ),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "Business/residential address (not a post box address)",
    ),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "City/town/suburb",
    ),
    ("Person/organisation sending the transfer instruction (if different)", "State"),
    ("Person/organisation sending the transfer instruction (if different)", "Postcode"),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "Postal address",
    ),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "City/town/suburb",
    ),
    ("Person/organisation sending the transfer instruction (if different)", "State"),
    ("Person/organisation sending the transfer instruction (if different)", "Postcode"),
    ("Person/organisation sending the transfer instruction (if different)", "Phone"),
    ("Person/organisation sending the transfer instruction (if different)", "Email"),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "Occupation, business or principal activity",
    ),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "ABN, ACN or ARBN",
    ),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "Business structure (if not an individual)",
    ),
    # Person/org receiving transfer instruction (8)
    ("Person/organisation receiving the transfer instruction", "Full name"),
    (
        "Person/organisation receiving the transfer instruction",
        "Business/residential address (not a post box address)",
    ),
    ("Person/organisation receiving the transfer instruction", "City/town/suburb"),
    ("Person/organisation receiving the transfer instruction", "State"),
    ("Person/organisation receiving the transfer instruction", "Postcode"),
    ("Person/organisation receiving the transfer instruction", "Country"),
    (
        "Person/organisation receiving the transfer instruction",
        "Is this person/organisation distributing money or property?",
    ),
    (
        "Person/organisation receiving the transfer instruction",
        "Is there a separate retail outlet/business location at which the money or property is being distributed?",
    ),
    # Person/org distributing (if different) (6)
    ("Person/organisation distributing money or property (if different)", "Full name"),
    (
        "Person/organisation distributing money or property (if different)",
        "Business/residential address (not a post box address)",
    ),
    (
        "Person/organisation distributing money or property (if different)",
        "City/town/suburb",
    ),
    ("Person/organisation distributing money or property (if different)", "State"),
    ("Person/organisation distributing money or property (if different)", "Postcode"),
    ("Person/organisation distributing money or property (if different)", "Country"),
    # Retail outlet (if different) (6)
    (
        "Retail outlet/business location where money or property is being distributed (if different)",
        "Full name",
    ),
    (
        "Retail outlet/business location where money or property is being distributed (if different)",
        "Business/residential address (not a post box address)",
    ),
    (
        "Retail outlet/business location where money or property is being distributed (if different)",
        "City/town/suburb",
    ),
    (
        "Retail outlet/business location where money or property is being distributed (if different)",
        "State",
    ),
    (
        "Retail outlet/business location where money or property is being distributed (if different)",
        "Postcode",
    ),
    (
        "Retail outlet/business location where money or property is being distributed (if different)",
        "Country",
    ),
    # Reason + Reporter (5)
    ("Reason", "Reason for the transfer"),
    ("Person completing this report", "Full name"),
    ("Person completing this report", "Job title"),
    ("Person completing this report", "Phone"),
    ("Person completing this report", "Email"),
]

IFTI_IN_COLUMNS: list[tuple[str, str]] = [
    # Transaction details (7)
    ("Transaction details", "Date money/property received from the ordering customer"),
    (
        "Transaction details",
        "Date money/property made available to the beneficiary customer",
    ),
    ("Transaction details", "Currency code"),
    ("Transaction details", "Total amount/value"),
    ("Transaction details", "Type of transfer"),
    ("Transaction details", "Description of property"),
    ("Transaction details", "Transaction reference number"),
    # Ordering customer (3)
    ("Ordering customer", "Full name"),
    ("Ordering customer", "If known by any other name"),
    ("Ordering customer", "Date of birth (if an individual)"),
    # Ordering customer contact details (12)
    (
        "Ordering customer contact details",
        "Business/residential address (not a post box address)",
    ),
    ("Ordering customer contact details", "City/town/suburb"),
    ("Ordering customer contact details", "State"),
    ("Ordering customer contact details", "Postcode"),
    ("Ordering customer contact details", "Country"),
    ("Ordering customer contact details", "Postal address"),
    ("Ordering customer contact details", "City/town/suburb"),
    ("Ordering customer contact details", "State"),
    ("Ordering customer contact details", "Postcode"),
    ("Ordering customer contact details", "Country"),
    ("Ordering customer contact details", "Phone"),
    ("Ordering customer contact details", "Email"),
    # Ordering customer business details (5 — no ID section for IN)
    (
        "Ordering customer business details",
        "Occupation, business or principal activity",
    ),
    ("Ordering customer business details", "ABN, ACN or ARBN"),
    ("Ordering customer business details", "Customer number (allocated by remitter)"),
    ("Ordering customer business details", "Account number"),
    ("Ordering customer business details", "Business structure (if not an individual)"),
    # Beneficiary customer (3)
    ("Beneficiary customer", "Full name"),
    ("Beneficiary customer", "Date of birth (if an individual)"),
    (
        "Beneficiary customer",
        "Any business name under which the beneficiary customer is operating",
    ),
    # Beneficiary customer contact details (12)
    (
        "Beneficiary customer contact details",
        "Business/residential address (not a post box address)",
    ),
    ("Beneficiary customer contact details", "City/town/suburb"),
    ("Beneficiary customer contact details", "State"),
    ("Beneficiary customer contact details", "Postcode"),
    ("Beneficiary customer contact details", "Country"),
    ("Beneficiary customer contact details", "Postal address"),
    ("Beneficiary customer contact details", "City/town/suburb"),
    ("Beneficiary customer contact details", "State"),
    ("Beneficiary customer contact details", "Postcode"),
    ("Beneficiary customer contact details", "Country"),
    ("Beneficiary customer contact details", "Phone"),
    ("Beneficiary customer contact details", "Email"),
    # Beneficiary customer business details (2)
    (
        "Beneficiary customer business details",
        "Occupation, business or principal activity",
    ),
    ("Beneficiary customer business details", "ABN, ACN or ARBN"),
    (
        "Beneficiary customer business details",
        "Business structure (if not an individual)",
    ),
    # Beneficiary customer account details (4)
    ("Beneficiary customer account details", "Account number"),
    (
        "Beneficiary customer account details",
        "Name of institution (where account is held)",
    ),
    ("Beneficiary customer account details", "City"),
    ("Beneficiary customer account details", "Country"),
    # Person/org accepting transfer instruction (19 cols for IN — has address, postal, phone, email, occ, etc.)
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Full name",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "If known by any other name",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Date of birth (if an individual)",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Business/residential address (not a post box address)",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "City/town/suburb",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "State",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Postcode",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Country",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Postal address",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "City/town/suburb",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "State",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Postcode",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Country",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Phone",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Email",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Occupation, business or principal activity",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Business structure (if not an individual)",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Is this person/organisation accepting the money or property?",
    ),
    (
        "Person/organisation accepting the transfer instruction from the ordering customer",
        "Is this person/organisation sending the transfer instruction?",
    ),
    # Person/org accepting money (if different) (6)
    (
        "Person/organisation accepting the money or property from the ordering customer (if different)",
        "Full name",
    ),
    (
        "Person/organisation accepting the money or property from the ordering customer (if different)",
        "Business/residential address (not a post box address)",
    ),
    (
        "Person/organisation accepting the money or property from the ordering customer (if different)",
        "City/town/suburb",
    ),
    (
        "Person/organisation accepting the money or property from the ordering customer (if different)",
        "State",
    ),
    (
        "Person/organisation accepting the money or property from the ordering customer (if different)",
        "Postcode",
    ),
    (
        "Person/organisation accepting the money or property from the ordering customer (if different)",
        "Country",
    ),
    # Person/org sending transfer instruction (if different) (18)
    (
        "Person/organisation sending the transfer instruction (if different)",
        "Full name",
    ),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "If known by any other name",
    ),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "Date of birth (if an individual)",
    ),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "Business/residential address (not a post box address)",
    ),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "City/town/suburb",
    ),
    ("Person/organisation sending the transfer instruction (if different)", "State"),
    ("Person/organisation sending the transfer instruction (if different)", "Postcode"),
    ("Person/organisation sending the transfer instruction (if different)", "Country"),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "Postal address",
    ),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "City/town/suburb",
    ),
    ("Person/organisation sending the transfer instruction (if different)", "State"),
    ("Person/organisation sending the transfer instruction (if different)", "Postcode"),
    ("Person/organisation sending the transfer instruction (if different)", "Country"),
    ("Person/organisation sending the transfer instruction (if different)", "Phone"),
    ("Person/organisation sending the transfer instruction (if different)", "Email"),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "Occupation, business or principal activity",
    ),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "ABN, ACN or ARBN",
    ),
    (
        "Person/organisation sending the transfer instruction (if different)",
        "Business structure (if not an individual)",
    ),
    # Person/org receiving transfer instruction (7)
    ("Person/organisation receiving the transfer instruction", "Full name"),
    (
        "Person/organisation receiving the transfer instruction",
        "Business/residential address (not a post box address)",
    ),
    ("Person/organisation receiving the transfer instruction", "City/town/suburb"),
    ("Person/organisation receiving the transfer instruction", "State"),
    ("Person/organisation receiving the transfer instruction", "Postcode"),
    (
        "Person/organisation receiving the transfer instruction",
        "Is this person/organisation distributing money or property?",
    ),
    (
        "Person/organisation receiving the transfer instruction",
        "Is there a separate retail outlet/business location at which the money or property is being distributed?",
    ),
    # Person/org distributing (if different) (5)
    ("Person/organisation distributing money or property (if different)", "Full name"),
    (
        "Person/organisation distributing money or property (if different)",
        "Business/residential address (not a post box address)",
    ),
    (
        "Person/organisation distributing money or property (if different)",
        "City/town/suburb",
    ),
    ("Person/organisation distributing money or property (if different)", "State"),
    ("Person/organisation distributing money or property (if different)", "Postcode"),
    # Retail outlet (6)
    (
        "Retail outlet/business location where money or property is being distributed (if different)",
        "Identification number of the retail outlet/business location",
    ),
    (
        "Retail outlet/business location where money or property is being distributed (if different)",
        "Full name",
    ),
    (
        "Retail outlet/business location where money or property is being distributed (if different)",
        "Business/residential address (not a post box address)",
    ),
    (
        "Retail outlet/business location where money or property is being distributed (if different)",
        "City/town/suburb",
    ),
    (
        "Retail outlet/business location where money or property is being distributed (if different)",
        "State",
    ),
    (
        "Retail outlet/business location where money or property is being distributed (if different)",
        "Postcode",
    ),
    # Reason + Reporter (5)
    ("Reason", "Reason for the transfer"),
    ("Person completing this report", "Full name"),
    ("Person completing this report", "Job title"),
    ("Person completing this report", "Phone"),
    ("Person completing this report", "Email"),
]


def _fmt_date(d) -> str:
    """Format date as DD/MM/YYYY string for AUSTRAC."""
    if d is None:
        return ""
    if isinstance(d, (date, datetime)):
        return d.strftime("%d/%m/%Y")
    return str(d)


def _row_out(r: IFTIRecord) -> list:
    """Map IFTIRecord → 112-value list matching IFTI-OUT column order."""
    return [
        _fmt_date(r.date_received),
        _fmt_date(r.date_available),
        r.currency_code or "AUD",
        r.total_amount or "",
        r.transfer_type or "Money",
        r.property_description or "",
        r.transaction_reference or "",
        # Ordering customer
        r.oc_full_name or "",
        r.oc_other_name or "",
        _fmt_date(r.oc_dob),
        r.oc_address or "",
        r.oc_city or "",
        r.oc_state or "",
        r.oc_postcode or "",
        r.oc_country or "",
        r.oc_postal_address or "",
        r.oc_postal_city or "",
        r.oc_postal_state or "",
        r.oc_postal_postcode or "",
        r.oc_postal_country or "",
        r.oc_phone or "",
        r.oc_email or "",
        r.oc_occupation or "",
        r.oc_abn or "",
        r.oc_customer_number or "",
        r.oc_account_number or "",
        r.oc_business_structure or "",
        # OC ID details
        r.oc_id1_type or "",
        r.oc_id1_type_other or "",
        r.oc_id1_number or "",
        r.oc_id1_issuer or "",
        r.oc_id2_type or "",
        r.oc_id2_type_other or "",
        r.oc_id2_number or "",
        r.oc_id2_issuer or "",
        r.oc_electronic_source or "",
        # Beneficiary customer
        r.bc_full_name or "",
        _fmt_date(r.bc_dob),
        r.bc_business_name or "",
        r.bc_address or "",
        r.bc_city or "",
        r.bc_state or "",
        r.bc_postcode or "",
        r.bc_country or "",
        r.bc_postal_address or "",
        r.bc_postal_city or "",
        r.bc_postal_state or "",
        r.bc_postal_postcode or "",
        r.bc_postal_country or "",
        r.bc_phone or "",
        r.bc_email or "",
        r.bc_occupation or "",
        r.bc_abn or "",
        r.bc_business_structure or "",
        r.bc_account_number or "",
        r.bc_institution_name or "",
        r.bc_institution_city or "",
        r.bc_institution_country or "",
        # Accept instruction block (OUT: retail ID first, then name/address/state/postcode, yes/no)
        r.retail_id_number or "",
        r.accept_full_name or "",
        r.accept_address or "",
        r.accept_city or "",
        r.accept_state or "",
        r.accept_postcode or "",
        r.is_accepting_money or "Yes",
        r.is_sending_instruction or "Yes",
        # Accepting money (if different)
        r.diff_accept_full_name or "",
        r.diff_accept_address or "",
        r.diff_accept_city or "",
        r.diff_accept_state or "",
        r.diff_accept_postcode or "",
        # Sending instruction (if different)
        r.send_full_name or "",
        r.send_other_name or "",
        _fmt_date(r.send_dob),
        r.send_address or "",
        r.send_city or "",
        r.send_state or "",
        r.send_postcode or "",
        r.send_postal_address or "",
        r.send_postal_city or "",
        r.send_postal_state or "",
        r.send_postal_postcode or "",
        r.send_phone or "",
        r.send_email or "",
        r.send_occupation or "",
        r.send_abn or "",
        r.send_business_structure or "",
        # Receiving transfer instruction
        r.recv_full_name or "",
        r.recv_address or "",
        r.recv_city or "",
        r.recv_state or "",
        r.recv_postcode or "",
        r.recv_country or "",
        r.is_distributing or "Yes",
        r.has_retail_outlet or "No",
        # Distributing (if different)
        r.dist_full_name or "",
        r.dist_address or "",
        r.dist_city or "",
        r.dist_state or "",
        r.dist_postcode or "",
        r.dist_country or "",
        # Retail outlet
        r.retail_full_name or "",
        r.retail_address or "",
        r.retail_city or "",
        r.retail_state or "",
        r.retail_postcode or "",
        r.retail_country or "",
        # Reason + reporter
        r.reason_for_transfer or "",
        r.reporter_full_name or "",
        r.reporter_job_title or "",
        r.reporter_phone or "",
        r.reporter_email or "",
    ]


def _row_in(r: IFTIRecord) -> list:
    """Map IFTIRecord → 115-value list matching IFTI-IN column order."""
    return [
        _fmt_date(r.date_received),
        _fmt_date(r.date_available),
        r.currency_code or "AUD",
        r.total_amount or "",
        r.transfer_type or "Money",
        r.property_description or "",
        r.transaction_reference or "",
        # Ordering customer
        r.oc_full_name or "",
        r.oc_other_name or "",
        _fmt_date(r.oc_dob),
        r.oc_address or "",
        r.oc_city or "",
        r.oc_state or "",
        r.oc_postcode or "",
        r.oc_country or "",
        r.oc_postal_address or "",
        r.oc_postal_city or "",
        r.oc_postal_state or "",
        r.oc_postal_postcode or "",
        r.oc_postal_country or "",
        r.oc_phone or "",
        r.oc_email or "",
        r.oc_occupation or "",
        r.oc_abn or "",
        r.oc_customer_number or "",
        r.oc_account_number or "",
        r.oc_business_structure or "",
        # Beneficiary customer (no ID section for IN)
        r.bc_full_name or "",
        _fmt_date(r.bc_dob),
        r.bc_business_name or "",
        r.bc_address or "",
        r.bc_city or "",
        r.bc_state or "",
        r.bc_postcode or "",
        r.bc_country or "",
        r.bc_postal_address or "",
        r.bc_postal_city or "",
        r.bc_postal_state or "",
        r.bc_postal_postcode or "",
        r.bc_postal_country or "",
        r.bc_phone or "",
        r.bc_email or "",
        r.bc_occupation or "",
        r.bc_abn or "",
        r.bc_business_structure or "",
        r.bc_account_number or "",
        r.bc_institution_name or "",
        r.bc_institution_city or "",
        r.bc_institution_country or "",
        # Accept instruction block (IN: full details + yes/no at end)
        r.accept_full_name or "",
        r.accept_other_name or "",
        _fmt_date(r.accept_dob),
        r.accept_address or "",
        r.accept_city or "",
        r.accept_state or "",
        r.accept_postcode or "",
        r.accept_country or "",
        r.accept_postal_address or "",
        r.accept_postal_city or "",
        r.accept_postal_state or "",
        r.accept_postal_postcode or "",
        r.accept_postal_country or "",
        r.accept_phone or "",
        r.accept_email or "",
        r.accept_occupation or "",
        r.accept_business_structure or "",
        r.is_accepting_money or "Yes",
        r.is_sending_instruction or "Yes",
        # Accepting money (if different) — IN has country
        r.diff_accept_full_name or "",
        r.diff_accept_address or "",
        r.diff_accept_city or "",
        r.diff_accept_state or "",
        r.diff_accept_postcode or "",
        r.diff_accept_country or "",
        # Sending instruction (if different) — IN has country fields
        r.send_full_name or "",
        r.send_other_name or "",
        _fmt_date(r.send_dob),
        r.send_address or "",
        r.send_city or "",
        r.send_state or "",
        r.send_postcode or "",
        r.send_country or "",
        r.send_postal_address or "",
        r.send_postal_city or "",
        r.send_postal_state or "",
        r.send_postal_postcode or "",
        r.send_postal_country or "",
        r.send_phone or "",
        r.send_email or "",
        r.send_occupation or "",
        r.send_abn or "",
        r.send_business_structure or "",
        # Receiving transfer instruction (IN: 7 cols, no country)
        r.recv_full_name or "",
        r.recv_address or "",
        r.recv_city or "",
        r.recv_state or "",
        r.recv_postcode or "",
        r.is_distributing or "Yes",
        r.has_retail_outlet or "No",
        # Distributing (if different)
        r.dist_full_name or "",
        r.dist_address or "",
        r.dist_city or "",
        r.dist_state or "",
        r.dist_postcode or "",
        # Retail outlet (IN: ID number first)
        r.retail_id_number or "",
        r.retail_full_name or "",
        r.retail_address or "",
        r.retail_city or "",
        r.retail_state or "",
        r.retail_postcode or "",
        # Reason + reporter
        r.reason_for_transfer or "",
        r.reporter_full_name or "",
        r.reporter_job_title or "",
        r.reporter_phone or "",
        r.reporter_email or "",
    ]


# ── Styles ────────────────────────────────────────────────────────────────────

_SECTION_FILL = PatternFill("solid", fgColor="4472C4")
_SECTION_FONT = Font(bold=True, color="FFFFFF", size=9)
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_HEADER_FONT = Font(bold=True, size=8)
_DATA_FONT = Font(size=9)
_BORDER_SIDE = Side(style="thin", color="B8CCE4")
_BORDER = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE, top=_BORDER_SIDE, bottom=_BORDER_SIDE
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def generate_ifti_excel(
    records: List[IFTIRecord],
    direction: str,
) -> bytes:
    """
    Generate AUSTRAC-compatible IFTI Excel workbook.

    Returns raw bytes (.xlsx) ready to stream or save.
    The sheet name and column structure match the official AUSTRAC template exactly.
    """
    is_out = direction == IFTIDirection.outgoing or direction == "outgoing"
    columns = IFTI_OUT_COLUMNS if is_out else IFTI_IN_COLUMNS
    sheet_name = "IFTI-DRA OUT" if is_out else "IFTI-DRA IN"
    row_mapper = _row_out if is_out else _row_in

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.showGridLines = False

    n_cols = len(columns)

    # ── Row 1: Section headers (merged across consecutive same-section columns) ──
    sections = [c[0] for c in columns]
    col = 1
    while col <= n_cols:
        section = sections[col - 1]
        span = 1
        while col + span - 1 < n_cols and sections[col + span - 1] == section:
            span += 1
        cell = ws.cell(row=1, column=col, value=section)
        cell.fill = _SECTION_FILL
        cell.font = _SECTION_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
        if span > 1:
            end_col = col + span - 1
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=end_col)
        col += span

    # ── Row 2: Column sub-headers ────────────────────────────────────────────
    for ci, (_, label) in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=ci, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER

    # ── Rows 3+: Data ─────────────────────────────────────────────────────────
    for ri, record in enumerate(records, start=3):
        row_data = row_mapper(record)
        for ci, value in enumerate(row_data, start=1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font = _DATA_FONT
            cell.alignment = _LEFT
            cell.border = _BORDER

    # ── Column widths ─────────────────────────────────────────────────────────
    for ci in range(1, n_cols + 1):
        col_letter = get_column_letter(ci)
        columns[ci - 1][0]
        # Narrow columns for simple fields, wider for name/address
        sub_label = columns[ci - 1][1].lower()
        if any(
            k in sub_label for k in ("address", "occupation", "reason", "name of inst")
        ):
            ws.column_dimensions[col_letter].width = 28
        elif any(k in sub_label for k in ("full name", "if known", "description")):
            ws.column_dimensions[col_letter].width = 22
        elif any(
            k in sub_label
            for k in (
                "date",
                "currency",
                "type",
                "postcode",
                "state",
                "phone",
                "email",
                "number",
                "abn",
            )
        ):
            ws.column_dimensions[col_letter].width = 16
        else:
            ws.column_dimensions[col_letter].width = 18

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 45
    ws.freeze_panes = "A3"

    # ── Instructions sheet ────────────────────────────────────────────────────
    wi = wb.create_sheet("Instructions")
    wi["A1"] = (
        "International Funds Transfer Instruction Report under a Designated Remittance Arrangement"
    )
    wi["A1"].font = Font(bold=True, size=11)
    wi["A3"] = (
        "Complete this form if you are a reporting entity receiving an international funds transfer "
        "instruction under a designated remittance arrangement as specified under items 3 and 4 of "
        "section 46 of the Anti-Money Laundering and Counter-Terrorism Financing Act 2006 (AML/CTF Act)."
    )
    wi["A5"] = (
        "Once complete, copy and paste the reports from this spreadsheet into the AUSTRAC Online "
        "spreadsheet and use the 'submit' button to send the reports to AUSTRAC."
    )
    wi["A7"] = (
        "This spreadsheet may only be used to report transactions which involved no more than a single "
        "ordering customer and beneficiary customer."
    )
    wi.column_dimensions["A"].width = 120
    for row in wi.iter_rows():
        for c in row:
            c.alignment = Alignment(wrap_text=True)

    # ── Reorder sheets ────────────────────────────────────────────────────────
    wb.move_sheet("Instructions", offset=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── DB helpers ────────────────────────────────────────────────────────────────


def list_ifti(
    db,
    industry_id: Optional[str] = None,
    direction: Optional[str] = None,
    status: Optional[str] = None,
    organisation_id: Optional[int] = None,
) -> list:
    from sqlalchemy import or_

    from app.models.ifti import IFTIRecord

    q = db.query(IFTIRecord)
    if organisation_id:
        q = q.filter(
            or_(
                IFTIRecord.organisation_id == organisation_id,
                (IFTIRecord.organisation_id.is_(None))
                & (IFTIRecord.industry_id == industry_id),
            )
        )
    elif industry_id:
        q = q.filter(IFTIRecord.industry_id == industry_id)
    if direction:
        q = q.filter(IFTIRecord.direction == direction)
    if status:
        q = q.filter(IFTIRecord.status == status)
    return q.order_by(IFTIRecord.date_received.desc()).all()


def get_ifti(db, ifti_id: str) -> Optional[IFTIRecord]:
    from app.models.ifti import IFTIRecord

    return db.query(IFTIRecord).filter(IFTIRecord.ifti_id == ifti_id).first()
