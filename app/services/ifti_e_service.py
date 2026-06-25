"""
IFTI-E (International Funds Transfer Instruction — E-currency / SWIFT) Service.

Generates AUSTRAC-compatible Excel spreadsheets matching the IFTI-E v1.3
official AUSTRAC template format.

Schema: IFTI-E v1.3 (IFTI-E-1-3.xsd)
Namespace: http://austrac.gov.au/schema/reporting/IFTI-E-1-3

Two modes:
  swift      — raw SWIFT message supplied; minimal structured data required
  structured — full party breakdown without raw SWIFT message

Party mapping (XML → model prefix):
  payer            → payer_*
  payerInstn       → payer_instn_*
  correspondentInstn → correspondent_instns (JSON array)
  payeeInstn       → payee_instn_*
  payee            → payee_*

Usage:
  workbook_bytes = generate_ifti_e_excel(records, direction="outgoing")
"""

from __future__ import annotations

import io
from datetime import date
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.ifti_e import IFTIEDirection, IFTIERecord

# ── Column definitions ────────────────────────────────────────────────────────
# Each entry: (section_label, column_label)

IFTI_E_COLUMNS: list[tuple[str, str]] = [
    # Transaction details (5)
    ("Transaction details", "Date instruction received/sent"),
    ("Transaction details", "Date funds made available"),
    ("Transaction details", "Currency code"),
    ("Transaction details", "Total amount"),
    ("Transaction details", "Transaction reference number"),
    # Payment purpose (2)
    ("Payment details", "Details of payment (purpose, max 140 chars)"),
    ("Payment details", "Sender to receiver information"),
    # Swift mode (2)
    ("SWIFT", "Same as SWIFT ordering customer (Yes/No)"),
    ("SWIFT", "Raw SWIFT message"),
    # Payer (3)
    ("Payer", "Full name"),
    ("Payer", "If known by any other name"),
    ("Payer", "Date of birth (if an individual)"),
    # Payer contact (10)
    ("Payer contact details", "Business/residential address"),
    ("Payer contact details", "City/town/suburb"),
    ("Payer contact details", "State"),
    ("Payer contact details", "Postcode"),
    ("Payer contact details", "Country"),
    ("Payer contact details", "Postal address"),
    ("Payer contact details", "City/town/suburb"),
    ("Payer contact details", "State"),
    ("Payer contact details", "Postcode"),
    ("Payer contact details", "Country"),
    # Payer business (5)
    ("Payer business details", "Phone"),
    ("Payer business details", "Email"),
    ("Payer business details", "Occupation/business activity"),
    ("Payer business details", "ABN, ACN or ARBN"),
    ("Payer business details", "Account number"),
    ("Payer business details", "Business structure (if not an individual)"),
    # Payer ID (6)
    ("Payer identification", "ID type (1)"),
    ("Payer identification", "Number"),
    ("Payer identification", "Issuer"),
    ("Payer identification", "ID type (2)"),
    ("Payer identification", "Number"),
    ("Payer identification", "Issuer"),
    ("Payer identification", "Electronic data source"),
    # Payer institution (5)
    ("Payer's institution", "Name"),
    ("Payer's institution", "SWIFT BIC / code"),
    ("Payer's institution", "Address"),
    ("Payer's institution", "City"),
    ("Payer's institution", "Country"),
    # Correspondent banks — up to 3 shown; full list in JSON (5 cols × 3)
    ("Correspondent institution 1", "Name"),
    ("Correspondent institution 1", "SWIFT BIC / code"),
    ("Correspondent institution 1", "Address"),
    ("Correspondent institution 1", "City"),
    ("Correspondent institution 1", "Country"),
    ("Correspondent institution 2", "Name"),
    ("Correspondent institution 2", "SWIFT BIC / code"),
    ("Correspondent institution 2", "Address"),
    ("Correspondent institution 2", "City"),
    ("Correspondent institution 2", "Country"),
    ("Correspondent institution 3", "Name"),
    ("Correspondent institution 3", "SWIFT BIC / code"),
    ("Correspondent institution 3", "Address"),
    ("Correspondent institution 3", "City"),
    ("Correspondent institution 3", "Country"),
    # Payee institution (5) — MANDATORY
    ("Payee's institution", "Name"),
    ("Payee's institution", "SWIFT BIC / code"),
    ("Payee's institution", "Address"),
    ("Payee's institution", "City"),
    ("Payee's institution", "Country"),
    # Payee (3)
    ("Payee (beneficiary)", "Full name"),
    ("Payee (beneficiary)", "Date of birth (if an individual)"),
    ("Payee (beneficiary)", "Any business name"),
    # Payee contact (8)
    ("Payee contact details", "Business/residential address"),
    ("Payee contact details", "City/town/suburb"),
    ("Payee contact details", "State"),
    ("Payee contact details", "Postcode"),
    ("Payee contact details", "Country"),
    ("Payee contact details", "Phone"),
    ("Payee contact details", "Email"),
    # Payee business (5)
    ("Payee business details", "Occupation/business activity"),
    ("Payee business details", "ABN, ACN or ARBN"),
    ("Payee business details", "Account number"),
    ("Payee business details", "IBAN"),
    ("Payee business details", "Business structure (if not an individual)"),
    # Reporter (4)
    ("Reporter", "Full name"),
    ("Reporter", "Job title"),
    ("Reporter", "Phone"),
    ("Reporter", "Email"),
]


def _fmt_date(d: Optional[date]) -> str:
    if d is None:
        return ""
    if isinstance(d, str):
        return d
    return d.strftime("%d/%m/%Y")


def _corr_field(record: IFTIERecord, idx: int, field: str) -> str:
    """Extract a field from correspondent_instns[idx]."""
    instns = record.correspondent_instns or []
    if idx >= len(instns):
        return ""
    return str(instns[idx].get(field, "") or "")


def _record_to_row(record: IFTIERecord) -> list:
    corr = record.correspondent_instns or []

    def _c(i: int, f: str) -> str:
        if i >= len(corr):
            return ""
        return str(corr[i].get(f, "") or "")

    return [
        # Transaction details
        _fmt_date(record.date_received),
        _fmt_date(record.date_available),
        record.currency_code or "AUD",
        record.total_amount or "",
        record.transaction_reference or "",
        # Payment details
        record.details_of_payment or "",
        record.sender_to_receiver_info or "",
        # Swift
        record.payer_same_as_swift_ord_cust or "",
        record.swift_msg or "",
        # Payer
        record.payer_full_name or "",
        record.payer_other_name or "",
        _fmt_date(record.payer_dob),
        # Payer contact
        record.payer_address or "",
        record.payer_city or "",
        record.payer_state or "",
        record.payer_postcode or "",
        record.payer_country or "",
        record.payer_postal_address or "",
        record.payer_postal_city or "",
        record.payer_postal_state or "",
        record.payer_postal_postcode or "",
        record.payer_postal_country or "",
        # Payer business
        record.payer_phone or "",
        record.payer_email or "",
        record.payer_occupation or "",
        " / ".join(
            filter(None, [record.payer_abn, record.payer_acn, record.payer_arbn])
        ),
        record.payer_account_number or "",
        record.payer_business_structure or "",
        # Payer ID
        record.payer_id1_type or "",
        record.payer_id1_number or "",
        record.payer_id1_issuer or "",
        record.payer_id2_type or "",
        record.payer_id2_number or "",
        record.payer_id2_issuer or "",
        record.payer_electronic_source or "",
        # Payer institution
        record.payer_instn_name or "",
        record.payer_instn_code or "",
        record.payer_instn_address or "",
        record.payer_instn_city or "",
        record.payer_instn_country or "",
        # Correspondent 1
        _c(0, "name"),
        _c(0, "code"),
        _c(0, "address"),
        _c(0, "city"),
        _c(0, "country"),
        # Correspondent 2
        _c(1, "name"),
        _c(1, "code"),
        _c(1, "address"),
        _c(1, "city"),
        _c(1, "country"),
        # Correspondent 3
        _c(2, "name"),
        _c(2, "code"),
        _c(2, "address"),
        _c(2, "city"),
        _c(2, "country"),
        # Payee institution
        record.payee_instn_name or "",
        record.payee_instn_code or "",
        record.payee_instn_address or "",
        record.payee_instn_city or "",
        record.payee_instn_country or "",
        # Payee
        record.payee_full_name or "",
        _fmt_date(record.payee_dob),
        record.payee_business_name or "",
        # Payee contact
        record.payee_address or "",
        record.payee_city or "",
        record.payee_state or "",
        record.payee_postcode or "",
        record.payee_country or "",
        record.payee_phone or "",
        record.payee_email or "",
        # Payee business
        record.payee_occupation or "",
        " / ".join(
            filter(None, [record.payee_abn, record.payee_acn, record.payee_arbn])
        ),
        record.payee_account_number or "",
        record.payee_account_iban or "",
        record.payee_business_structure or "",
        # Reporter
        record.reporter_full_name or "",
        record.reporter_job_title or "",
        record.reporter_phone or "",
        record.reporter_email or "",
    ]


# ── Excel styles ──────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_SUBHEAD_FILL = PatternFill("solid", fgColor="2E75B6")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
_SUBHEAD_FONT = Font(bold=True, color="FFFFFF", size=8)
_DATA_FONT = Font(size=9)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def generate_ifti_e_excel(
    records: List[IFTIERecord],
    direction: IFTIEDirection = IFTIEDirection.outgoing,
) -> bytes:
    """
    Render an IFTI-E workbook matching the AUSTRAC IFTI-E v1.3 Excel template.
    Returns raw .xlsx bytes suitable for streaming.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"IFTI-E {direction.value.upper()}"

    # Row 1 — Section headers (merged per section group)
    # Row 2 — Column sub-headers
    sections: list[str] = [s for s, _ in IFTI_E_COLUMNS]
    labels: list[str] = [c for _, c in IFTI_E_COLUMNS]

    # Write section headers with merging
    col = 1
    run_start = 1
    run_label = sections[0]
    for i, sec in enumerate(sections):
        ws.cell(row=1, column=i + 1).value = (
            sec if sec != sections[i - 1] or i == 0 else None
        )
        ws.cell(row=1, column=i + 1).fill = _HEADER_FILL
        ws.cell(row=1, column=i + 1).font = _HEADER_FONT
        ws.cell(row=1, column=i + 1).alignment = _CENTER
        ws.cell(row=1, column=i + 1).border = _BORDER

    # Merge consecutive identical section headers
    col = 1
    while col <= len(sections):
        run_end = col
        while run_end < len(sections) and sections[run_end] == sections[col - 1]:
            run_end += 1
        if run_end - col > 0:
            ws.merge_cells(
                start_row=1,
                start_column=col,
                end_row=1,
                end_column=run_end,
            )
            ws.cell(row=1, column=col).value = sections[col - 1]
        col = run_end + 1

    # Row 2 — column sub-headers
    for i, label in enumerate(labels):
        cell = ws.cell(row=2, column=i + 1, value=label)
        cell.fill = _SUBHEAD_FILL
        cell.font = _SUBHEAD_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER

    # Data rows
    for row_idx, record in enumerate(records, start=3):
        for col_idx, value in enumerate(_record_to_row(record), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _DATA_FONT
            cell.alignment = _LEFT
            cell.border = _BORDER

    # Column widths
    for i in range(1, len(IFTI_E_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 45
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── DB helpers ────────────────────────────────────────────────────────────────


def list_ifti_e(
    db,
    org_id: str,
    direction: Optional[IFTIEDirection] = None,
    status: Optional[str] = None,
) -> List[IFTIERecord]:
    from app.models.ifti_e import IFTIERecord

    q = db.query(IFTIERecord).filter(IFTIERecord.industry_id == org_id)
    if direction:
        q = q.filter(IFTIERecord.direction == direction)
    if status:
        q = q.filter(IFTIERecord.status == status)
    return q.order_by(IFTIERecord.created_at.desc()).all()


def get_ifti_e(db, record_id: int, org_id: str) -> Optional[IFTIERecord]:
    from app.models.ifti_e import IFTIERecord

    return (
        db.query(IFTIERecord)
        .filter(
            IFTIERecord.id == record_id,
            IFTIERecord.industry_id == org_id,
        )
        .first()
    )
